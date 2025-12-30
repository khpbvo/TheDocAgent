"""XLSX spreadsheet editor with diff preview and approval workflow."""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .base import (
    ApprovalTracker,
    DocumentEditor,
    DocumentOperation,
    OperationResult,
    OperationType,
)
from .diff_display import render_cell_diff, DIM, CYAN, RESET


@dataclass
class XlsxOperation(DocumentOperation):
    """XLSX-specific operation with cell addressing."""
    
    # For range operations
    start_row: int | None = None
    end_row: int | None = None
    start_col: int | None = None
    end_col: int | None = None
    
    # For formula operations
    formula: str | None = None


class XlsxEditor(DocumentEditor):
    """Editor for XLSX files with cell change preview."""
    
    def get_supported_extensions(self) -> list[str]:
        return [".xlsx", ".xlsm", ".xls"]
    
    def render_diff(self, operation: DocumentOperation) -> str:
        """Render a human-readable diff for XLSX operations."""
        if operation.cell_range:
            # Range operation - show summary
            return self._render_range_diff(operation)
        elif operation.cell:
            # Single cell operation
            return render_cell_diff(
                cell=operation.cell,
                old_value=operation.old_value,
                new_value=operation.new_value,
                sheet=operation.sheet,
            )
        else:
            return f"{DIM}(no cell specified){RESET}"
    
    def _render_range_diff(self, operation: DocumentOperation) -> str:
        """Render diff for a range operation."""
        lines = []
        location = f"{operation.sheet}!{operation.cell_range}" if operation.sheet else operation.cell_range
        lines.append(f"{CYAN}Range: {location}{RESET}")
        
        if isinstance(operation.old_value, list) and isinstance(operation.new_value, list):
            # Show row-by-row diff for small ranges
            old_rows = operation.old_value
            new_rows = operation.new_value
            
            max_preview = 5
            for i, (old_row, new_row) in enumerate(zip(old_rows[:max_preview], new_rows[:max_preview])):
                if old_row != new_row:
                    lines.append(f"  Row {i+1}:")
                    lines.append(f"    \033[91m- {old_row}\033[0m")
                    lines.append(f"    \033[92m+ {new_row}\033[0m")
            
            if len(old_rows) > max_preview or len(new_rows) > max_preview:
                lines.append(f"  {DIM}... and more rows{RESET}")
        else:
            lines.append(render_cell_diff(
                cell=operation.cell_range or "range",
                old_value=str(operation.old_value),
                new_value=str(operation.new_value),
                sheet=operation.sheet,
            ))
        
        return "\n".join(lines)
    
    def apply_operation(self, operation: DocumentOperation) -> OperationResult:
        """Apply an XLSX operation."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return OperationResult(
                success=False,
                output="openpyxl not installed. Run: pip install openpyxl",
            )
        
        path = self._resolve(operation.path)
        
        if operation.type == OperationType.XLSX_WRITE_CELL:
            return self._write_cell(path, operation)
        elif operation.type == OperationType.XLSX_WRITE_RANGE:
            return self._write_range(path, operation)
        elif operation.type == OperationType.XLSX_ADD_FORMULA:
            return self._add_formula(path, operation)
        elif operation.type == OperationType.XLSX_DELETE_ROW:
            return self._delete_row(path, operation)
        elif operation.type == OperationType.XLSX_INSERT_ROW:
            return self._insert_row(path, operation)
        else:
            return OperationResult(
                success=False,
                output=f"Unsupported operation type: {operation.type}",
            )
    
    def _write_cell(self, path: Path, operation: DocumentOperation) -> OperationResult:
        """Write a value to a cell."""
        from openpyxl import load_workbook
        
        if not path.exists():
            return OperationResult(success=False, output=f"File not found: {path}")
        
        wb = load_workbook(str(path))
        
        # Get the sheet
        sheet_name = operation.sheet or wb.active.title
        if sheet_name not in wb.sheetnames:
            return OperationResult(
                success=False,
                output=f"Sheet not found: {sheet_name}",
            )
        
        ws = wb[sheet_name]
        cell_ref = operation.cell
        
        if not cell_ref:
            return OperationResult(success=False, output="No cell reference provided")
        
        # Write the value
        ws[cell_ref] = operation.new_value
        
        wb.save(str(path))
        return OperationResult(
            success=True,
            output=f"Updated {sheet_name}!{cell_ref} = {operation.new_value}",
        )
    
    def _write_range(self, path: Path, operation: DocumentOperation) -> OperationResult:
        """Write values to a range of cells."""
        from openpyxl import load_workbook
        
        if not path.exists():
            return OperationResult(success=False, output=f"File not found: {path}")
        
        if not isinstance(operation.new_value, list):
            return OperationResult(
                success=False,
                output="Range operation requires list of values",
            )
        
        wb = load_workbook(str(path))
        sheet_name = operation.sheet or wb.active.title
        
        if sheet_name not in wb.sheetnames:
            return OperationResult(success=False, output=f"Sheet not found: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # Parse the range or use start position
        if isinstance(operation, XlsxOperation) and operation.start_row is not None:
            start_row = operation.start_row
            start_col = operation.start_col or 1
        elif operation.cell:
            # Use cell as starting point
            from openpyxl.utils import coordinate_from_string, column_index_from_string
            col_letter, row = coordinate_from_string(operation.cell)
            start_row = row
            start_col = column_index_from_string(col_letter)
        else:
            start_row = 1
            start_col = 1
        
        # Write the values
        for row_idx, row_data in enumerate(operation.new_value):
            if isinstance(row_data, list):
                for col_idx, value in enumerate(row_data):
                    ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
            else:
                ws.cell(row=start_row + row_idx, column=start_col, value=row_data)
        
        wb.save(str(path))
        return OperationResult(
            success=True,
            output=f"Updated range in {sheet_name}",
        )
    
    def _add_formula(self, path: Path, operation: DocumentOperation) -> OperationResult:
        """Add a formula to a cell."""
        from openpyxl import load_workbook
        
        if not path.exists():
            return OperationResult(success=False, output=f"File not found: {path}")
        
        wb = load_workbook(str(path))
        sheet_name = operation.sheet or wb.active.title
        
        if sheet_name not in wb.sheetnames:
            return OperationResult(success=False, output=f"Sheet not found: {sheet_name}")
        
        ws = wb[sheet_name]
        cell_ref = operation.cell
        
        if not cell_ref:
            return OperationResult(success=False, output="No cell reference provided")
        
        formula = operation.new_value
        if isinstance(operation, XlsxOperation) and operation.formula:
            formula = operation.formula
        
        if not str(formula).startswith("="):
            formula = f"={formula}"
        
        ws[cell_ref] = formula
        
        wb.save(str(path))
        return OperationResult(
            success=True,
            output=f"Added formula to {sheet_name}!{cell_ref}: {formula}",
        )
    
    def _delete_row(self, path: Path, operation: DocumentOperation) -> OperationResult:
        """Delete a row from the spreadsheet."""
        from openpyxl import load_workbook
        
        if not path.exists():
            return OperationResult(success=False, output=f"File not found: {path}")
        
        wb = load_workbook(str(path))
        sheet_name = operation.sheet or wb.active.title
        
        if sheet_name not in wb.sheetnames:
            return OperationResult(success=False, output=f"Sheet not found: {sheet_name}")
        
        ws = wb[sheet_name]
        
        if isinstance(operation, XlsxOperation) and operation.start_row is not None:
            row_idx = operation.start_row
        else:
            return OperationResult(success=False, output="No row index provided")
        
        ws.delete_rows(row_idx)
        
        wb.save(str(path))
        return OperationResult(
            success=True,
            output=f"Deleted row {row_idx} from {sheet_name}",
        )
    
    def _insert_row(self, path: Path, operation: DocumentOperation) -> OperationResult:
        """Insert a row into the spreadsheet."""
        from openpyxl import load_workbook
        
        if not path.exists():
            return OperationResult(success=False, output=f"File not found: {path}")
        
        wb = load_workbook(str(path))
        sheet_name = operation.sheet or wb.active.title
        
        if sheet_name not in wb.sheetnames:
            return OperationResult(success=False, output=f"Sheet not found: {sheet_name}")
        
        ws = wb[sheet_name]
        
        if isinstance(operation, XlsxOperation) and operation.start_row is not None:
            row_idx = operation.start_row
        else:
            return OperationResult(success=False, output="No row index provided")
        
        ws.insert_rows(row_idx)
        
        # If we have new values, write them
        if operation.new_value and isinstance(operation.new_value, list):
            for col_idx, value in enumerate(operation.new_value, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(str(path))
        return OperationResult(
            success=True,
            output=f"Inserted row at position {row_idx} in {sheet_name}",
        )
    
    # Helper methods for creating operations
    
    @classmethod
    def create_cell_operation(
        cls,
        path: str,
        cell: str,
        new_value: Any,
        old_value: Any = None,
        sheet: str | None = None,
        description: str = "",
    ) -> XlsxOperation:
        """Create a cell write operation."""
        return XlsxOperation(
            type=OperationType.XLSX_WRITE_CELL,
            path=path,
            description=description or f"Update cell {cell}",
            cell=cell,
            sheet=sheet,
            old_value=old_value,
            new_value=new_value,
        )
    
    @classmethod
    def create_formula_operation(
        cls,
        path: str,
        cell: str,
        formula: str,
        old_value: Any = None,
        sheet: str | None = None,
        description: str = "",
    ) -> XlsxOperation:
        """Create a formula operation."""
        return XlsxOperation(
            type=OperationType.XLSX_ADD_FORMULA,
            path=path,
            description=description or f"Add formula to {cell}",
            cell=cell,
            sheet=sheet,
            old_value=old_value,
            new_value=formula,
            formula=formula,
        )
    
    @classmethod
    def create_range_operation(
        cls,
        path: str,
        cell_range: str,
        new_values: list,
        old_values: list | None = None,
        sheet: str | None = None,
        description: str = "",
    ) -> XlsxOperation:
        """Create a range write operation."""
        return XlsxOperation(
            type=OperationType.XLSX_WRITE_RANGE,
            path=path,
            description=description or f"Update range {cell_range}",
            cell_range=cell_range,
            sheet=sheet,
            old_value=old_values,
            new_value=new_values,
        )
