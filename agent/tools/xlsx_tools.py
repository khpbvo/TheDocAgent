"""XLSX spreadsheet analysis and modification tools.

Wraps existing skills/xlsx/ utilities for use with OpenAI Agents SDK.
"""

import json
import sys
from pathlib import Path

from agents import function_tool

from .output_utils import truncate_json_output, truncate_output

# Add skills directory to path for imports
SKILLS_DIR = Path(__file__).parent.parent.parent / "skills"
sys.path.insert(0, str(SKILLS_DIR / "xlsx"))


@function_tool
def get_sheet_names(file_path: str) -> str:
    """Get the names of all sheets in an Excel file.

    Args:
        file_path: Path to the Excel file (.xlsx or .xlsm).

    Returns:
        List of sheet names in the workbook.
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return f"Sheets in workbook ({len(sheets)} total): {', '.join(sheets)}"
    except Exception as e:
        return f"Error reading Excel file: {e!s}"


@function_tool
def read_sheet(
    file_path: str,
    sheet_name: str | None = None,
    start_row: int = 1,
    max_rows: int = 100,
) -> str:
    """Read data from an Excel sheet with pagination support.

    For large sheets, use start_row and max_rows to paginate through data,
    or use search_sheet() to find specific content first.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to read. If not provided, reads the active sheet.
        start_row: Starting row number (1-indexed, default: 1). Use for pagination.
        max_rows: Maximum number of rows to return (default: 100).

    Returns:
        JSON string containing the sheet data as a list of rows with pagination info.
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    try:
        # data_only=True to get computed values, not formulas
        wb = load_workbook(file_path, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Count total rows
        total_rows = ws.max_row or 0

        data = []
        rows_processed = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # Skip rows before start_row
            if row_idx < start_row:
                continue
            # Stop after max_rows
            if rows_processed >= max_rows:
                break
            # Convert None to empty string for cleaner output
            data.append([str(cell) if cell is not None else "" for cell in row])
            rows_processed += 1

        wb.close()

        end_row = start_row + len(data) - 1 if data else start_row
        has_more = end_row < total_rows

        result = {
            "sheet_name": ws.title,
            "start_row": start_row,
            "end_row": end_row,
            "rows_returned": len(data),
            "total_rows": total_rows,
            "has_more_rows": has_more,
            "data": data,
        }

        if has_more:
            result["next_start_row"] = end_row + 1
            result["tip"] = f"Use start_row={end_row + 1} to get next page"

        return truncate_json_output(json.dumps(result, indent=2))
    except Exception as e:
        return f"Error reading Excel sheet: {e!s}"


@function_tool
def get_formulas(file_path: str, sheet_name: str | None = None) -> str:
    """Get formulas from an Excel sheet (not computed values).

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to read. If not provided, reads the active sheet.

    Returns:
        JSON string containing cells with formulas and their formula strings.
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    try:
        # data_only=False to get formulas, not values
        wb = load_workbook(file_path, data_only=False)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return f"Error: Sheet '{sheet_name}' not found."
            ws = wb[sheet_name]
        else:
            ws = wb.active

        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})

        wb.close()

        if not formulas:
            return f"No formulas found in sheet '{ws.title}'."

        return json.dumps(
            {
                "sheet_name": ws.title,
                "formula_count": len(formulas),
                "formulas": formulas,
            },
            indent=2,
        )
    except Exception as e:
        return f"Error reading formulas: {e!s}"


@function_tool
def analyze_data(
    file_path: str,
    sheet_name: str | None = None,
    analysis_type: str = "summary",
) -> str:
    """Perform data analysis on an Excel sheet using pandas.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to analyze. If not provided, uses the first sheet.
        analysis_type: Type of analysis to perform:
            - "summary": Basic statistics (describe())
            - "info": Data types and null counts
            - "head": First 10 rows
            - "shape": Row and column counts

    Returns:
        Analysis results as formatted text.
    """
    import pandas as pd

    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        if analysis_type == "summary":
            result = f"Statistical Summary for '{sheet_name or 'Sheet1'}':\n{df.describe().to_string()}"
            return truncate_output(result)
        if analysis_type == "info":
            info_str = f"Shape: {df.shape[0]} rows x {df.shape[1]} columns\n\n"
            info_str += "Column Info:\n"
            for col in df.columns:
                non_null = df[col].notna().sum()
                info_str += (
                    f"  {col}: {df[col].dtype}, {non_null}/{len(df)} non-null values\n"
                )
            return truncate_output(info_str)
        if analysis_type == "head":
            result = f"First 10 rows:\n{df.head(10).to_string()}"
            return truncate_output(result)
        if analysis_type == "shape":
            return f"Dataset shape: {df.shape[0]} rows, {df.shape[1]} columns\nColumns: {', '.join(str(c) for c in df.columns)}"
        return f"Unknown analysis type: {analysis_type}. Use 'summary', 'info', 'head', or 'shape'."
    except Exception as e:
        return f"Error analyzing Excel data: {e!s}"


@function_tool
def write_cell(
    file_path: str,
    sheet_name: str,
    cell: str,
    value: str,
    output_path: str | None = None,
) -> str:
    """Write a value to a specific cell in an Excel file.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to modify.
        cell: Cell reference (e.g., "A1", "B5").
        value: Value to write to the cell.
        output_path: Where to save. If not provided, overwrites the input file.

    Returns:
        Success message or error.
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    try:
        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."

        ws = wb[sheet_name]
        ws[cell] = value

        save_path = output_path or file_path
        wb.save(save_path)
        wb.close()

        return f"Successfully wrote '{value}' to {sheet_name}!{cell}. Saved to: {save_path}"
    except Exception as e:
        return f"Error writing to cell: {e!s}"


@function_tool
def add_formula(
    file_path: str,
    sheet_name: str,
    cell: str,
    formula: str,
    output_path: str | None = None,
) -> str:
    """Add an Excel formula to a specific cell.

    Note: The formula will not be calculated until the file is opened in Excel
    or recalculated using recalculate_formulas().

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to modify.
        cell: Cell reference (e.g., "A1", "B5").
        formula: Excel formula (must start with "=", e.g., "=SUM(A1:A10)").
        output_path: Where to save. If not provided, overwrites the input file.

    Returns:
        Success message or error.
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    if not formula.startswith("="):
        formula = "=" + formula

    try:
        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."

        ws = wb[sheet_name]
        ws[cell] = formula

        save_path = output_path or file_path
        wb.save(save_path)
        wb.close()

        return f"Successfully added formula '{formula}' to {sheet_name}!{cell}. Saved to: {save_path}. Note: Run recalculate_formulas() to compute the value."
    except Exception as e:
        return f"Error adding formula: {e!s}"


@function_tool
def recalculate_formulas(file_path: str, timeout: int = 30) -> str:
    """Recalculate all Excel formulas using LibreOffice.

    Directly uses skills/xlsx/recalc.py

    Args:
        file_path: Path to the Excel file.
        timeout: Maximum time to wait for recalculation in seconds (default 30).

    Returns:
        JSON string with recalculation results:
        - status: 'success' or 'errors_found'
        - total_errors: Number of formula errors found
        - total_formulas: Number of formulas in the file
        - error_summary: Breakdown by error type (#REF!, #DIV/0!, etc.)
    """
    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    try:
        # Import existing skill function
        from recalc import recalc

        result = recalc(file_path, timeout)
        return json.dumps(result, indent=2)
    except ImportError:
        return "Error: recalc module not found. Make sure skills/xlsx/recalc.py exists and LibreOffice is installed."
    except Exception as e:
        return f"Error recalculating formulas: {e!s}"


@function_tool
def search_sheet(
    file_path: str,
    query: str,
    sheet_name: str | None = None,
    case_sensitive: bool = False,
    max_results: int = 50,
) -> str:
    """Search for text within an Excel sheet and return matching cells.

    Use this to find specific content in large spreadsheets before reading full data.
    Returns cell references and values for each match.

    Args:
        file_path: Path to the Excel file to search.
        query: Text to search for in cell values.
        sheet_name: Name of sheet to search. If not provided, searches active sheet.
        case_sensitive: Whether the search should be case-sensitive (default: False).
        max_results: Maximum number of results to return (default: 50).

    Returns:
        JSON with matching cells, their values, and row numbers.
        Use the row numbers with read_sheet(start_row=N) to get surrounding data.
    """
    from openpyxl import load_workbook

    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    if not query or not query.strip():
        return "Error: Search query cannot be empty."

    try:
        wb = load_workbook(file_path, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            ws = wb[sheet_name]
        else:
            ws = wb.active

        results = []
        total_matches = 0
        rows_with_matches = set()
        search_query = query if case_sensitive else query.lower()

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                cell_str = str(cell.value)
                search_text = cell_str if case_sensitive else cell_str.lower()

                if search_query in search_text:
                    total_matches += 1
                    rows_with_matches.add(cell.row)

                    if len(results) < max_results:
                        results.append(
                            {
                                "cell": cell.coordinate,
                                "row": cell.row,
                                "column": cell.column,
                                "value": (
                                    cell_str[:200] + "..."
                                    if len(cell_str) > 200
                                    else cell_str
                                ),
                            }
                        )

        wb.close()

        if not results:
            return f"No matches found for '{query}' in the sheet."

        output = {
            "query": query,
            "sheet_name": ws.title,
            "total_matches": total_matches,
            "rows_with_matches": sorted(rows_with_matches)[:50],  # Limit row list
            "results": results,
            "tip": "Use read_sheet(start_row=N) with these row numbers to get surrounding data.",
        }

        return json.dumps(output, indent=2)
    except Exception as e:
        return f"Error searching Excel sheet: {e!s}"
