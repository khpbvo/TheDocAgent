"""Document modification tools with diff preview and approval.

These tools show a diff and ask for confirmation before modifying documents.
They wrap the editors from agent.editors for use with the OpenAI Agents SDK.
"""

import os
from pathlib import Path
from typing import Any

from agents import function_tool

# Get workspace root from environment or use current directory
def _get_workspace_root() -> Path:
    """Get the workspace root directory."""
    root = os.environ.get("DOCUMENT_WORKSPACE_ROOT")
    if root:
        return Path(root).resolve()
    return Path.cwd().resolve()


# Lazy initialization of shared state
_approval_tracker = None
_docx_editor = None
_xlsx_editor = None
_auto_approve = False


def _get_auto_approve() -> bool:
    """Check if auto-approve is enabled."""
    return os.environ.get("DOCUMENT_EDIT_AUTO_APPROVE") == "1"


def _get_approval_tracker():
    """Get or create the shared approval tracker."""
    global _approval_tracker
    if _approval_tracker is None:
        from ..editors.base import ApprovalTracker
        _approval_tracker = ApprovalTracker()
    return _approval_tracker


def _get_docx_editor():
    """Get or create the DOCX editor."""
    global _docx_editor
    if _docx_editor is None:
        from ..editors.docx_editor import DocxEditor
        _docx_editor = DocxEditor(
            root=_get_workspace_root(),
            approvals=_get_approval_tracker(),
            auto_approve=_get_auto_approve(),
        )
    return _docx_editor


def _get_xlsx_editor():
    """Get or create the XLSX editor."""
    global _xlsx_editor
    if _xlsx_editor is None:
        from ..editors.xlsx_editor import XlsxEditor
        _xlsx_editor = XlsxEditor(
            root=_get_workspace_root(),
            approvals=_get_approval_tracker(),
            auto_approve=_get_auto_approve(),
        )
    return _xlsx_editor


def reset_approval_state():
    """Reset all approval state. Call when starting a new session."""
    global _approval_tracker, _docx_editor, _xlsx_editor
    if _approval_tracker:
        _approval_tracker.clear()
    _docx_editor = None
    _xlsx_editor = None


def set_workspace_root(path: str | Path):
    """Set the workspace root directory."""
    global _docx_editor, _xlsx_editor
    os.environ["DOCUMENT_WORKSPACE_ROOT"] = str(Path(path).resolve())
    # Reset editors so they pick up new root
    _docx_editor = None
    _xlsx_editor = None


# =============================================================================
# DOCX Tools with Approval
# =============================================================================

@function_tool
def replace_docx_text(
    file_path: str,
    old_text: str,
    new_text: str,
    description: str = "",
) -> str:
    """Replace text in a Word document with diff preview and approval.
    
    Shows a diff of the proposed change and asks for confirmation before applying.
    
    Args:
        file_path: Path to the DOCX file.
        old_text: The exact text to find and replace.
        new_text: The replacement text.
        description: Optional description of why this change is being made.
    
    Returns:
        Result message indicating success or failure.
    """
    from ..editors.docx_editor import DocxEditor
    
    editor = _get_docx_editor()
    
    # Try to get context around the old text
    context_before = None
    context_after = None
    location = None
    
    try:
        from docx import Document
        path = Path(file_path)
        if path.exists():
            doc = Document(str(path))
            for i, para in enumerate(doc.paragraphs):
                if old_text in para.text:
                    location = f"Paragraph {i + 1}"
                    # Get surrounding context
                    text = para.text
                    idx = text.find(old_text)
                    if idx > 0:
                        context_before = text[max(0, idx-50):idx].strip()
                    if idx + len(old_text) < len(text):
                        context_after = text[idx+len(old_text):idx+len(old_text)+50].strip()
                    break
    except Exception:
        pass  # Context extraction is best-effort
    
    operation = DocxEditor.create_replace_operation(
        path=file_path,
        old_text=old_text,
        new_text=new_text,
        description=description,
        location=location,
        context_before=context_before,
        context_after=context_after,
    )
    
    result = editor.execute(operation)
    return result.output


@function_tool
def insert_docx_text(
    file_path: str,
    new_text: str,
    paragraph_index: int = -1,
    description: str = "",
) -> str:
    """Insert text into a Word document with diff preview and approval.
    
    Shows the proposed insertion and asks for confirmation before applying.
    
    Args:
        file_path: Path to the DOCX file.
        new_text: The text to insert.
        paragraph_index: Where to insert (-1 = end of document, 0 = beginning).
        description: Optional description of why this change is being made.
    
    Returns:
        Result message indicating success or failure.
    """
    from ..editors.docx_editor import DocxEditor, DocxOperation
    from ..editors.base import OperationType
    
    editor = _get_docx_editor()
    
    location = "End of document" if paragraph_index == -1 else f"After paragraph {paragraph_index}"
    
    operation = DocxOperation(
        type=OperationType.DOCX_INSERT_TEXT,
        path=file_path,
        description=description or "Insert text",
        new_text=new_text,
        location=location,
        paragraph_index=None if paragraph_index == -1 else paragraph_index,
    )
    
    result = editor.execute(operation)
    return result.output


@function_tool
def delete_docx_text(
    file_path: str,
    text_to_delete: str,
    description: str = "",
) -> str:
    """Delete text from a Word document with diff preview and approval.
    
    Shows what will be deleted and asks for confirmation before applying.
    
    Args:
        file_path: Path to the DOCX file.
        text_to_delete: The exact text to find and delete.
        description: Optional description of why this deletion is needed.
    
    Returns:
        Result message indicating success or failure.
    """
    from ..editors.docx_editor import DocxEditor
    
    editor = _get_docx_editor()
    
    operation = DocxEditor.create_delete_operation(
        path=file_path,
        old_text=text_to_delete,
        description=description,
    )
    
    result = editor.execute(operation)
    return result.output


# =============================================================================
# XLSX Tools with Approval
# =============================================================================

@function_tool
def update_xlsx_cell(
    file_path: str,
    cell: str,
    new_value: str,
    sheet: str = "",
    description: str = "",
) -> str:
    """Update a cell in an Excel spreadsheet with diff preview and approval.
    
    Shows the old and new values and asks for confirmation before applying.
    
    Args:
        file_path: Path to the XLSX file.
        cell: Cell reference (e.g., "A1", "B5").
        new_value: The new value to write.
        sheet: Sheet name (uses active sheet if not specified).
        description: Optional description of why this change is being made.
    
    Returns:
        Result message indicating success or failure.
    """
    from ..editors.xlsx_editor import XlsxEditor
    
    editor = _get_xlsx_editor()
    
    # Get current value for diff display
    old_value = None
    try:
        from openpyxl import load_workbook
        path = Path(file_path)
        if path.exists():
            wb = load_workbook(str(path), data_only=True)
            sheet_name = sheet or wb.active.title
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                old_value = ws[cell].value
            wb.close()
    except Exception:
        pass  # Old value extraction is best-effort
    
    operation = XlsxEditor.create_cell_operation(
        path=file_path,
        cell=cell,
        new_value=new_value,
        old_value=old_value,
        sheet=sheet or None,
        description=description,
    )
    
    result = editor.execute(operation)
    return result.output


@function_tool
def add_xlsx_formula(
    file_path: str,
    cell: str,
    formula: str,
    sheet: str = "",
    description: str = "",
) -> str:
    """Add a formula to an Excel cell with diff preview and approval.
    
    Shows the formula and asks for confirmation before applying.
    
    Args:
        file_path: Path to the XLSX file.
        cell: Cell reference (e.g., "A1", "B5").
        formula: The formula to add (with or without leading =).
        sheet: Sheet name (uses active sheet if not specified).
        description: Optional description of why this formula is needed.
    
    Returns:
        Result message indicating success or failure.
    """
    from ..editors.xlsx_editor import XlsxEditor
    
    editor = _get_xlsx_editor()
    
    # Get current value for diff display
    old_value = None
    try:
        from openpyxl import load_workbook
        path = Path(file_path)
        if path.exists():
            wb = load_workbook(str(path))
            sheet_name = sheet or wb.active.title
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                old_value = ws[cell].value
            wb.close()
    except Exception:
        pass
    
    operation = XlsxEditor.create_formula_operation(
        path=file_path,
        cell=cell,
        formula=formula,
        old_value=old_value,
        sheet=sheet or None,
        description=description,
    )
    
    result = editor.execute(operation)
    return result.output


@function_tool
def update_xlsx_range(
    file_path: str,
    start_cell: str,
    values_json: str,
    sheet: str = "",
    description: str = "",
) -> str:
    """Update a range of cells in an Excel spreadsheet with diff preview.
    
    Shows proposed changes and asks for confirmation before applying.
    
    Args:
        file_path: Path to the XLSX file.
        start_cell: Top-left cell of the range (e.g., "A1").
        values_json: JSON array of arrays with new values.
                     Example: '[["A","B"],["C","D"]]' for a 2x2 range.
        sheet: Sheet name (uses active sheet if not specified).
        description: Optional description of why these changes are needed.
    
    Returns:
        Result message indicating success or failure.
    """
    import json
    from ..editors.xlsx_editor import XlsxEditor
    
    try:
        values = json.loads(values_json)
    except json.JSONDecodeError as e:
        return f"Error: Invalid JSON for values: {e}"
    
    if not isinstance(values, list):
        return "Error: values must be a JSON array"
    
    editor = _get_xlsx_editor()
    
    # Calculate range from start cell and values dimensions
    rows = len(values)
    cols = max(len(row) if isinstance(row, list) else 1 for row in values) if values else 1
    
    # Parse start cell to get end cell
    from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
    col_letter, row = coordinate_from_string(start_cell)
    start_col = column_index_from_string(col_letter)
    end_col = start_col + cols - 1
    end_row = row + rows - 1
    cell_range = f"{start_cell}:{get_column_letter(end_col)}{end_row}"
    
    operation = XlsxEditor.create_range_operation(
        path=file_path,
        cell_range=cell_range,
        new_values=values,
        sheet=sheet or None,
        description=description,
    )
    
    result = editor.execute(operation)
    return result.output


# Export all approval-enabled tools
APPROVAL_TOOLS = [
    replace_docx_text,
    insert_docx_text,
    delete_docx_text,
    update_xlsx_cell,
    add_xlsx_formula,
    update_xlsx_range,
]

__all__ = [
    "replace_docx_text",
    "insert_docx_text",
    "delete_docx_text",
    "update_xlsx_cell",
    "add_xlsx_formula",
    "update_xlsx_range",
    "APPROVAL_TOOLS",
    "reset_approval_state",
    "set_workspace_root",
]
