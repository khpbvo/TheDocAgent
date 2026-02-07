"""Directed search and retrieval tools for large documents.

These tools are designed to keep context usage low by:
1) searching and ranking small text segments, then
2) retrieving only the selected segments.
"""

import heapq
import json
import re
import zipfile
from pathlib import Path

from agents import function_tool

from .output_utils import truncate_json_output

SEGMENT_MAX_CHARS = 1200
SEGMENT_OVERLAP_CHARS = 200
MAX_TOP_K = 50
MAX_RETRIEVE_CHARS = 30000


def _split_text_segments(
    text: str,
    max_chars: int = SEGMENT_MAX_CHARS,
    overlap_chars: int = SEGMENT_OVERLAP_CHARS,
) -> list[str]:
    """Split text into overlapping chunks for ranked retrieval."""
    cleaned = " ".join(text.split())
    if not cleaned:
        return []
    if len(cleaned) <= max_chars:
        return [cleaned]

    segments: list[str] = []
    start = 0
    while start < len(cleaned):
        end = min(start + max_chars, len(cleaned))
        segments.append(cleaned[start:end].strip())
        if end >= len(cleaned):
            break
        start = max(0, end - overlap_chars)
    return segments


def _query_terms(query: str, case_sensitive: bool) -> list[str]:
    if case_sensitive:
        terms = re.findall(r"\w+", query)
    else:
        terms = re.findall(r"\w+", query.lower())
    return [term for term in terms if term.strip()]


def _score_text(
    text: str,
    query: str,
    mode: str,
    case_sensitive: bool,
) -> tuple[float, int, int, int]:
    """Return score plus match statistics."""
    haystack = text if case_sensitive else text.lower()
    needle = query if case_sensitive else query.lower()
    terms = _query_terms(query, case_sensitive)

    phrase_hits = haystack.count(needle) if needle else 0
    term_hits = sum(haystack.count(term) for term in terms)
    unique_hits = sum(1 for term in terms if term in haystack)

    if mode == "exact":
        score = phrase_hits * 25.0
    elif mode == "terms":
        score = unique_hits * 6.0 + term_hits * 1.5
    else:  # hybrid
        density = (term_hits / max(len(haystack), 1)) * 1000.0
        score = (
            phrase_hits * 20.0
            + unique_hits * 5.0
            + term_hits * 1.0
            + min(density, 5.0)
        )

    return score, phrase_hits, term_hits, unique_hits


def _build_snippet(
    text: str,
    query: str,
    case_sensitive: bool,
    context_chars: int,
) -> str:
    haystack = text if case_sensitive else text.lower()
    needle = query if case_sensitive else query.lower()
    pos = haystack.find(needle) if needle else -1

    if pos == -1:
        snippet = text[: min(len(text), context_chars * 2)].strip()
        return snippet + ("..." if len(text) > len(snippet) else "")

    start = max(0, pos - context_chars)
    end = min(len(text), pos + len(query) + context_chars)
    snippet = text[start:end].strip()
    if start > 0:
        snippet = "..." + snippet
    if end < len(text):
        snippet = snippet + "..."
    return snippet


def _iter_pdf_segments(file_path: Path):
    import pdfplumber

    with pdfplumber.open(file_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            for segment_idx, segment_text in enumerate(_split_text_segments(text), 1):
                yield {
                    "selector": {"page": page_num, "segment": segment_idx},
                    "location": f"page {page_num}, segment {segment_idx}",
                    "text": segment_text,
                }


def _iter_docx_segments(file_path: Path):
    from defusedxml import ElementTree as ET

    with zipfile.ZipFile(file_path, "r") as zf:
        document_xml = zf.read("word/document.xml")

    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    root = ET.fromstring(document_xml)

    paragraph_index = 0
    for para in root.findall(".//w:p", ns):
        text_elements = para.findall(".//w:t", ns)
        para_text = "".join(t.text or "" for t in text_elements).strip()
        if not para_text:
            continue

        paragraph_index += 1
        for segment_idx, segment_text in enumerate(_split_text_segments(para_text), 1):
            yield {
                "selector": {"paragraph": paragraph_index, "segment": segment_idx},
                "location": f"paragraph {paragraph_index}, segment {segment_idx}",
                "text": segment_text,
            }


def _iter_xlsx_segments(file_path: Path, sheet_name: str | None):
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        if sheet_name and sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {available}"
            )

        sheet_names = [sheet_name] if sheet_name else wb.sheetnames
        for current_sheet in sheet_names:
            ws = wb[current_sheet]
            for row in ws.iter_rows():
                row_parts = []
                for cell in row:
                    if cell.value is None:
                        continue
                    value = str(cell.value).strip()
                    if not value:
                        continue
                    compact = value if len(value) <= 180 else value[:180] + "..."
                    row_parts.append(f"{cell.coordinate}: {compact}")

                if not row_parts:
                    continue

                row_text = " | ".join(row_parts)
                for segment_idx, segment_text in enumerate(
                    _split_text_segments(row_text),
                    1,
                ):
                    yield {
                        "selector": {
                            "sheet": current_sheet,
                            "row": row[0].row if row else 0,
                            "segment": segment_idx,
                        },
                        "location": f"sheet '{current_sheet}', row {row[0].row if row else 0}, segment {segment_idx}",
                        "text": segment_text,
                    }
    finally:
        wb.close()


def _detect_file_type(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return "pdf"
    if suffix == ".docx":
        return "docx"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    return "unknown"


@function_tool
def directed_search_document(
    file_path: str,
    query: str,
    top_k: int = 8,
    mode: str = "hybrid",
    case_sensitive: bool = False,
    context_chars: int = 120,
    sheet_name: str | None = None,
) -> str:
    """Rank search hits in a large document without returning full document text.

    Use this before extraction on large files. It returns compact hits with
    selectors that can be passed to retrieve_document_segments().

    Args:
        file_path: Path to PDF, DOCX, XLSX, or XLSM file.
        query: Search text.
        top_k: Maximum number of ranked hits to return (1-50).
        mode: Ranking mode: "hybrid" (default), "exact", or "terms".
        case_sensitive: Whether matching is case-sensitive.
        context_chars: Context width around the first match in each hit snippet.
        sheet_name: Optional sheet restriction for Excel files.

    Returns:
        JSON with ranked hits and selector objects.
    """
    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"
    if not query or not query.strip():
        return "Error: Search query cannot be empty."

    normalized_mode = mode.lower().strip()
    if normalized_mode not in {"hybrid", "exact", "terms"}:
        return "Error: mode must be one of: hybrid, exact, terms."

    k = max(1, min(top_k, MAX_TOP_K))
    file_type = _detect_file_type(path)
    if file_type == "unknown":
        return f"Error: Unsupported file type: {path.suffix}. Supported: .pdf, .docx, .xlsx, .xlsm"

    if file_type == "pdf":
        iterator = _iter_pdf_segments(path)
    elif file_type == "docx":
        iterator = _iter_docx_segments(path)
    else:
        iterator = _iter_xlsx_segments(path, sheet_name)

    heap: list[tuple[float, int, dict]] = []
    seen_segments = 0
    matched_segments = 0
    counter = 0

    try:
        for segment in iterator:
            seen_segments += 1
            score, phrase_hits, term_hits, unique_hits = _score_text(
                text=segment["text"],
                query=query,
                mode=normalized_mode,
                case_sensitive=case_sensitive,
            )
            if score <= 0:
                continue

            matched_segments += 1
            entry = {
                "score": round(score, 4),
                "selector": segment["selector"],
                "location": segment["location"],
                "snippet": _build_snippet(
                    text=segment["text"],
                    query=query,
                    case_sensitive=case_sensitive,
                    context_chars=max(20, context_chars),
                ),
                "stats": {
                    "phrase_hits": phrase_hits,
                    "term_hits": term_hits,
                    "unique_terms_matched": unique_hits,
                },
            }

            counter += 1
            if len(heap) < k:
                heapq.heappush(heap, (score, counter, entry))
            elif score > heap[0][0]:
                heapq.heapreplace(heap, (score, counter, entry))
    except Exception as e:
        return f"Error searching document: {e!s}"

    ranked_hits = [item[2] for item in sorted(heap, key=lambda item: item[0], reverse=True)]
    output = {
        "query": query,
        "file_path": str(path),
        "file_type": file_type,
        "mode": normalized_mode,
        "top_k": k,
        "segments_scanned": seen_segments,
        "segments_matched": matched_segments,
        "hits": ranked_hits,
        "tip": "Pass one or more hit selectors to retrieve_document_segments(selectors_json='[...]') for focused extraction.",
    }
    return truncate_json_output(json.dumps(output, indent=2))


def _extract_selectors(selectors_json: str) -> list[dict]:
    raw = json.loads(selectors_json)
    if not isinstance(raw, list):
        raise ValueError("selectors_json must be a JSON array.")

    selectors: list[dict] = []
    for item in raw:
        if isinstance(item, dict) and isinstance(item.get("selector"), dict):
            selectors.append(item["selector"])
        elif isinstance(item, dict):
            selectors.append(item)
    return selectors


def _match_pdf_segment(segment: dict, selector: dict, neighborhood: int) -> bool:
    page = int(selector.get("page", 0))
    seg = int(selector.get("segment", 0)) if selector.get("segment") else None
    current_page = int(segment["selector"]["page"])
    current_seg = int(segment["selector"]["segment"])

    if neighborhood <= 0:
        if seg is None:
            return current_page == page
        return current_page == page and current_seg == seg

    if abs(current_page - page) > neighborhood:
        return False
    if seg is None:
        return True
    if current_page != page:
        return True
    return abs(current_seg - seg) <= neighborhood


def _match_docx_segment(segment: dict, selector: dict, neighborhood: int) -> bool:
    paragraph = int(selector.get("paragraph", 0))
    seg = int(selector.get("segment", 0)) if selector.get("segment") else None
    current_paragraph = int(segment["selector"]["paragraph"])
    current_seg = int(segment["selector"]["segment"])

    if neighborhood <= 0:
        if seg is None:
            return current_paragraph == paragraph
        return current_paragraph == paragraph and current_seg == seg

    if abs(current_paragraph - paragraph) > neighborhood:
        return False
    if seg is None:
        return True
    if current_paragraph != paragraph:
        return True
    return abs(current_seg - seg) <= neighborhood


def _match_xlsx_segment(segment: dict, selector: dict, neighborhood: int) -> bool:
    sheet = selector.get("sheet")
    row = int(selector.get("row", 0))
    seg = int(selector.get("segment", 0)) if selector.get("segment") else None

    current_sheet = str(segment["selector"]["sheet"])
    current_row = int(segment["selector"]["row"])
    current_seg = int(segment["selector"]["segment"])

    if sheet and current_sheet != sheet:
        return False

    if neighborhood <= 0:
        if seg is None:
            return current_row == row
        return current_row == row and current_seg == seg

    if abs(current_row - row) > neighborhood:
        return False
    if seg is None:
        return True
    if current_row != row:
        return True
    return abs(current_seg - seg) <= neighborhood


@function_tool
def retrieve_document_segments(
    file_path: str,
    selectors_json: str,
    neighborhood: int = 0,
    max_chars: int = 8000,
    sheet_name: str | None = None,
) -> str:
    """Retrieve only selected segments from a document using selectors.

    This tool is meant to follow directed_search_document(). It keeps context
    usage predictable by returning only exact segments (plus optional local
    neighborhood).

    Args:
        file_path: Path to PDF, DOCX, XLSX, or XLSM file.
        selectors_json: JSON array of selector objects. You may pass raw
            selectors or full hit objects from directed_search_document().
        neighborhood: Optional surrounding neighborhood (0-3) around each
            selector (pages/paragraphs/rows depending on file type).
        max_chars: Approximate output budget for serialized JSON (1000-30000).
        sheet_name: Optional sheet restriction for Excel files.

    Returns:
        JSON containing retrieved segments with location metadata.
    """
    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    try:
        selectors = _extract_selectors(selectors_json)
    except Exception as e:
        return f"Error: Invalid selectors_json: {e!s}"

    if not selectors:
        return "Error: selectors_json must include at least one selector object."

    bounded_neighborhood = max(0, min(neighborhood, 3))
    budget = max(1000, min(max_chars, MAX_RETRIEVE_CHARS))
    file_type = _detect_file_type(path)
    if file_type == "unknown":
        return f"Error: Unsupported file type: {path.suffix}. Supported: .pdf, .docx, .xlsx, .xlsm"

    try:
        if file_type == "pdf":
            all_segments = list(_iter_pdf_segments(path))
            matcher = _match_pdf_segment
        elif file_type == "docx":
            all_segments = list(_iter_docx_segments(path))
            matcher = _match_docx_segment
        else:
            all_segments = list(_iter_xlsx_segments(path, sheet_name))
            matcher = _match_xlsx_segment
    except Exception as e:
        return f"Error retrieving segments: {e!s}"

    results: list[dict] = []
    seen_keys: set[str] = set()
    content_chars = 0
    truncated = False

    for selector in selectors:
        if not isinstance(selector, dict):
            continue
        for segment in all_segments:
            try:
                if not matcher(segment, selector, bounded_neighborhood):
                    continue
            except (TypeError, ValueError):
                continue

            key = json.dumps(segment["selector"], sort_keys=True)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            segment_text = segment["text"]
            remaining = budget - content_chars
            if remaining < 200:
                truncated = True
                break
            if len(segment_text) > remaining:
                segment_text = segment_text[:remaining].rstrip() + "..."
                truncated = True

            results.append(
                {
                    "selector": segment["selector"],
                    "location": segment["location"],
                    "text": segment_text,
                }
            )
            content_chars += len(segment_text)
        if truncated:
            break

    output = {
        "file_path": str(path),
        "file_type": file_type,
        "selectors_requested": len(selectors),
        "segments_returned": len(results),
        "neighborhood": bounded_neighborhood,
        "truncated": truncated,
        "results": results,
    }
    return truncate_json_output(json.dumps(output, indent=2), max_chars=budget)
